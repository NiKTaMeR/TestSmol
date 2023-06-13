import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime

def parse_information(raw_data):
    # Find earliest and latest transaction dates
    earliest_date = raw_data['Transaction Date'].min()
    latest_date = raw_data['Transaction Date'].max()

    # Create a new Excel file with "Inputs_Variables" sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs_Variables"

    ws['A1'] = "threshold_contraction"
    ws['B1'] = 0.1  # Default percentage value, user can modify it later
    ws['A2'] = "threshold_upsell"
    ws['B2'] = 0.1  # Default percentage value, user can modify it later

    # Prompt user for path and filename
    file_path = input("Enter the path and filename for the new Excel file (e.g., C:/Users/username/Documents/new_file.xlsx): ")

    # Save the new Excel file
    wb.save(file_path)

    # Open the new Excel file for the user to fill in the monthly results
    os.startfile(file_path)

    # Wait for the user to finish editing the file and press the button
    input("Now opening the excel file, fill in the monthly results, save and close the file, before coming back here and press Enter.")

    # Read the updated Excel file
    updated_wb = pd.read_excel(file_path, sheet_name="Inputs_Variables")

    # Get the threshold values
    threshold_contraction = updated_wb.loc[0, 'B']
    threshold_upsell = updated_wb.loc[1, 'B']

    return earliest_date, latest_date, threshold_contraction, threshold_upsell